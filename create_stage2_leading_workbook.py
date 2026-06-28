from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
REPORTS_DIR = BASE_DIR / "reports"
INDUSTRY_CSV = REPORTS_DIR / "screener_industry_weinstein_rrg_2026-06-28.csv"
STOCK_CSV = REPORTS_DIR / "screener_industry_stock_rankings_2026-06-28.csv"
OUTPUT = REPORTS_DIR / "stage2_rrg_leading_industries_best_stocks_2026-06-28.xlsx"


def value_or_blank(value):
    if pd.isna(value):
        return ""
    return value


def main() -> None:
    industries = pd.read_csv(INDUSTRY_CSV)
    stocks = pd.read_csv(STOCK_CSV)

    leading = industries[
        (industries["stage"] == "stage_2")
        & (industries["rrg_quadrant"] == "leading")
    ].copy()
    leading = leading.sort_values(
        ["rs_momentum", "rs_ratio", "rs_ratio_5d_delta"],
        ascending=[False, False, False],
    )

    rows = []
    for _, industry in leading.iterrows():
        ranked = stocks[stocks["parent"] == industry["name"]].copy()
        ranked = ranked.sort_values(
            ["stock_score", "stage2_score", "rs_momentum", "market_cap_cr"],
            ascending=[False, False, False, False],
        ).head(5)
        for rank, (_, stock) in enumerate(ranked.iterrows(), start=1):
            rows.append(
                {
                    "Industry": industry["name"],
                    "Sector": industry["parent"],
                    "NSE Industry": industry["nse_industry"],
                    "Industry Members Used": industry["member_count"],
                    "Industry Stage": industry["stage"],
                    "Industry Stage2 Score": industry["stage2_score"],
                    "Industry RRG 5D Ago": industry["rrg_quadrant_5d_ago"],
                    "Industry RRG Now": industry["rrg_quadrant"],
                    "Industry RS Ratio": industry["rs_ratio"],
                    "Industry RS Momentum": industry["rs_momentum"],
                    "Industry 5D RS Ratio Delta": industry["rs_ratio_5d_delta"],
                    "Industry 5D RS Momentum Delta": industry["rs_momentum_5d_delta"],
                    "Just Entered Leading": industry["just_entered_leading"],
                    "Stock Rank": rank,
                    "Stock": stock["name"],
                    "Symbol": stock["symbol"],
                    "Stock Stage": stock["stage"],
                    "Stock Stage2 Score": stock["stage2_score"],
                    "Stock RRG 5D Ago": stock["rrg_quadrant_5d_ago"],
                    "Stock RRG Now": stock["rrg_quadrant"],
                    "Stock RS Ratio": stock["rs_ratio"],
                    "Stock RS Momentum": stock["rs_momentum"],
                    "Stock 30W MA Slope %": stock["weekly_ma_slope_4w_pct"],
                    "Market Cap Cr": stock["market_cap_cr"],
                    "P/E": stock["pe"],
                    "ROCE %": stock["roce"],
                    "Technical Score": stock["stock_score"],
                    "Industry Source": industry["source_url"],
                }
            )

    result = pd.DataFrame(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stage2 RRG Leading"

    ws["A1"] = "Weinstein Stage 2 + RRG Leading Industries: Best Stocks"
    ws["A2"] = f"Generated: {date.today().isoformat()}"
    ws["A3"] = (
        "Coverage note: this uses the latest available Screener/Yahoo run. Screener rate-limited the full scrape, "
        "so the current source analysis covered 44 industries from the reference workbook."
    )
    ws["A4"] = "Stock ranking basis: Stage 2 score, RRG quadrant, RS Ratio/Momentum, 30-week MA slope, and market-cap metadata."

    for row in range(1, 5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(result.columns))
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A2"].font = Font(italic=True)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")

    header_row = 6
    for col_idx, column in enumerate(result.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(result.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value_or_blank(value))
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(result.columns))}{header_row + len(result)}"

    percent_like = {"Industry 5D RS Ratio Delta", "Industry 5D RS Momentum Delta", "Stock 30W MA Slope %", "ROCE %"}
    two_decimals = {
        "Industry RS Ratio",
        "Industry RS Momentum",
        "Stock RS Ratio",
        "Stock RS Momentum",
        "Market Cap Cr",
        "P/E",
        "Technical Score",
    }
    for col_idx, column in enumerate(result.columns, start=1):
        letter = get_column_letter(col_idx)
        if column in percent_like or column in two_decimals:
            for row_idx in range(header_row + 1, header_row + 1 + len(result)):
                ws[f"{letter}{row_idx}"].number_format = "0.00"

    widths = {
        "A": 30,
        "B": 28,
        "C": 24,
        "N": 10,
        "O": 24,
        "P": 18,
        "AB": 55,
    }
    for col_idx in range(1, len(result.columns) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(letter, 14)

    for row in range(1, header_row + 1 + len(result)):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 40
    ws.row_dimensions[4].height = 34

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
