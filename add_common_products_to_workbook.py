from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
REPORTS_DIR = BASE_DIR / "reports"
INPUT = REPORTS_DIR / "stage2_rrg_leading_industries_best_stocks_2026-06-28.xlsx"
OUTPUT = REPORTS_DIR / "stage2_rrg_leading_industries_best_stocks_products_2026-06-28.xlsx"


PRODUCTS = {
    "Fusion Finance": "Micro loans / joint-liability group loans for women entrepreneurs",
    "Muthoot Microfin": "Microfinance loans for women and small borrowers",
    "CreditAcc. Gram.": "Grameen-style microfinance loans",
    "Satin Creditcare": "Microfinance and small business loans",
    "Spandana Sphoort": "Microfinance loans for rural women borrowers",
    "HFCL": "Optical fibre cable, telecom network equipment, Wi-Fi/5G gear",
    "Suyog Telematics": "Telecom towers and network infrastructure leasing",
    "Pace Digitek": "Telecom and digital infrastructure services",
    "GTL Infra.": "Telecom tower infrastructure",
    "Indus Towers": "Mobile telecom towers used by Indian telecom operators",
    "SKM Egg Prod.": "Egg powder and processed egg products",
    "Venky's (India)": "Venky's chicken, poultry products, eggs and animal feed",
    "HMA Agro Inds.": "Frozen buffalo meat and processed meat exports",
    "Bharat Forge": "Forged auto and industrial components",
    "Samvardh. Mothe.": "Motherson auto wiring harnesses, mirrors and vehicle modules",
    "Sona BLW Precis.": "EV and auto drivetrain parts such as gears and motors",
    "Bosch": "Bosch auto parts, fuel injection systems, spark plugs and power tools",
    "ZF Commercial": "Braking, suspension and safety systems for commercial vehicles",
    "Grindwell Norton": "Norton abrasives, grinding wheels and industrial ceramics",
    "Carborundum Uni.": "CUMI abrasives, grinding wheels and industrial ceramics",
    "Timken India": "Timken bearings for vehicles and industrial machines",
    "Wendt India": "Super abrasives and precision grinding tools",
    "SKF India Indus.": "SKF bearings for industrial machinery",
    "Ador Welding": "Welding electrodes, welding machines and consumables",
    "Inox India": "Cryogenic tanks and LNG/industrial gas storage equipment",
    "Diffusion Eng": "Welding consumables and wear-protection engineering products",
    "HBL Engineering": "Industrial batteries, defence electronics and railway electronics",
    "Harsha Engg Intl": "Bearing cages and precision engineering components",
    "ESAF Small Fin": "ESAF small finance banking and microloans",
    "AU Small Finance": "AU Small Finance Bank savings accounts, loans and cards",
    "Equitas Sma. Fin": "Equitas Small Finance Bank accounts and vehicle/MSME loans",
    "Suryoday Small": "Suryoday Small Finance Bank deposits and micro/MSME loans",
    "Capital Small": "Capital Small Finance Bank deposits and loans",
    "R R Kabel": "RR Kabel electric wires and cables",
    "Quadrant Future": "Railway signalling cables and specialty cables",
    "Paramount Comm.": "Paramount electric and telecom cables",
    "Finolex Cables": "Finolex electrical wires, cables and PVC conduits",
    "V-Marc India": "V-Marc wires and electrical cables",
    "Sigma Advanced System": "Aerospace and defence electronics / systems",
    "Astra Microwave": "Defence radar, microwave and RF systems",
    "Data Pattern": "Defence electronics, radars and avionics systems",
    "Apollo Micro Sys": "Defence and aerospace electronic systems",
    "Zen Technologies": "Defence simulators, anti-drone systems and training equipment",
    "Nippon Life Ind.": "Nippon India Mutual Fund schemes",
    "Aditya AMC": "Aditya Birla Sun Life Mutual Fund schemes",
    "IL&FS Inv.Manag.": "Private equity and infrastructure fund management",
    "HDFC AMC": "HDFC Mutual Fund schemes",
    "ICICI AMC": "ICICI Prudential Mutual Fund schemes",
    "Prec. Wires (I)": "Copper winding wires used in motors and transformers",
    "Baheti Recycling": "Aluminium recycling and aluminium alloy products",
    "Shera Energy": "Copper, brass and aluminium winding wires/strips",
    "M Tek Copper": "Copper wires, rods and copper products",
    "Vidya Wires": "Copper winding wires for electrical equipment",
}


def main() -> None:
    wb = load_workbook(INPUT)
    ws = wb["Stage2 RRG Leading"]
    header_row = 6
    stock_col = None
    for cell in ws[header_row]:
        if cell.value == "Stock":
            stock_col = cell.column
            break
    if stock_col is None:
        raise RuntimeError("Stock column not found.")

    insert_at = stock_col + 1
    ws.insert_cols(insert_at)
    product_header = ws.cell(row=header_row, column=insert_at, value="Commonly Known Product / Business")
    product_header.font = Font(bold=True, color="FFFFFF")
    product_header.fill = PatternFill("solid", fgColor="5B9BD5")
    product_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_row = ws.max_row
    for row_idx in range(header_row + 1, max_row + 1):
        stock = ws.cell(row=row_idx, column=stock_col).value
        product = PRODUCTS.get(stock, "")
        cell = ws.cell(row=row_idx, column=insert_at, value=product)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions[get_column_letter(insert_at)].width = 42
    for row_idx in range(header_row + 1, max_row + 1):
        ws.row_dimensions[row_idx].height = 34

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws["A4"] = (
        f"{ws['A4'].value} Product/business labels are plain-English descriptions of what the company is commonly associated with in India."
    )
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
